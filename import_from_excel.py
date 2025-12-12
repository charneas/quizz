import openpyxl
from sqlmodel import Session, select
from app.db import engine
from app.models import Question
from datetime import datetime

def map_difficulty(difficulty_str):
    """Convertit la difficulté en nombre"""
    if not difficulty_str:
        return 1
    diff_lower = str(difficulty_str).lower()
    if 'facile' in diff_lower:
        return 1
    elif 'moyen' in diff_lower:
        return 2
    elif 'difficile' in diff_lower:
        return 3
    else:
        return 1

def import_questions_from_excel():
    """Importe les questions depuis le fichier Excel"""
    
    # Charger le fichier Excel
    workbook = openpyxl.load_workbook('Questions Themes.xlsx')
    
    # Parcourir chaque feuille
    with Session(engine) as session:
        imported_count = 0
        
        # Traiter la feuille "Questions"
        if 'Questions' in workbook.sheetnames:
            print("\n=== Traitement de la feuille: Questions ===")
            sheet = workbook['Questions']
            
            # Parcourir les lignes (en commençant à la ligne 2)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Si la première colonne est vide, on saute
                    continue
                
                # Format: Difficulté, Thème, Question, Réponse
                difficulty_str = row[0]
                theme = row[1]
                question_text = row[2]
                answer_text = row[3]
                
                if not question_text:
                    continue
                
                try:
                    # Éviter les doublons
                    existing = session.exec(
                        select(Question).where(Question.text == question_text)
                    ).first()
                    
                    if not existing:
                        question = Question(
                            text=str(question_text),
                            answer=str(answer_text) if answer_text else "Non défini",
                            theme=str(theme) if theme else "Général",
                            difficulty=map_difficulty(difficulty_str),
                            points=10,
                            created_at=datetime.utcnow(),
                            created_by="Excel import"
                        )
                        session.add(question)
                        imported_count += 1
                        print(f"  ✓ {question_text[:50]}... ({theme})")
                except Exception as e:
                    print(f"  ✗ Erreur: {e}")
        
        
        # Commit final
        try:
            session.commit()
            print(f"\n✅ {imported_count} questions importées avec succès!")
        except Exception as e:
            session.rollback()
            print(f"✗ Erreur lors du commit: {e}")

if __name__ == "__main__":
    import_questions_from_excel()
